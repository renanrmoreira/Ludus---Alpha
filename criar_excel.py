from openpyxl import Workbook, load_workbook
import os

def adicionar_dados_excel(nome_arquivo, data_dicionario):

        caminho_arquivo = os.path.join(os.getcwd(), nome_arquivo)
        
        if os.path.exists(caminho_arquivo):
            wb = load_workbook(caminho_arquivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(data_dicionario.keys()))
        
        ws.append(list(data_dicionario.values()))
        
        wb.save(caminho_arquivo)
